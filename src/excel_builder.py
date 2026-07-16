"""
excel_builder.py
----------------
Builds the final, formatted Excel workbook from parsed records: one row
per source document, with a two-row header (parameter name spanning
Min/Target/Max sub-columns), matching the layout used in prior manual
deliverables. Columns are whatever the caller selects (e.g. via the
API selected_columns list), so a file lacking a given parameter simply gets
blank cells in that row for it.
"""

from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .classifier import SINGLE_VALUE_COLUMNS
from .extractor import clean_param_value

IDENTITY_HEADERS = [
    ("SpecNo", "Spec. No."),
    ("Client", "Client"),
    ("Quality", "Quality"),
    ("Grade", "Grade"),
    ("MatCode", "Mat. Code"),
    ("Color", "Color"),
    ("Ply", "Ply"),
]

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
DATA_FILL = PatternFill("solid", fgColor="E2EFDA")
HEADER_FONT = Font(name="Arial", size=10, bold=True)
DATA_FONT = Font(name="Calibri", size=11)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def build_workbook(records, selected_columns, sort_by="SpecNo"):
    """
    Build and return an openpyxl Workbook.

    records: list of record dicts from extractor.parse_folder
    selected_columns: ordered list of canonical parameter names to include
        (e.g. from the API selected_columns list) -- a subset of, or equal to,
        the full dynamically-discovered column set
    sort_by: record key to sort rows by (falls back to filename as a
        secondary key so duplicate spec numbers stay in a stable order)
    """
    records_sorted = sorted(records, key=lambda r: (r.get(sort_by, ""), r.get("file", "")))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Specifications"

    # ---- header: row 1 (group labels) + row 2 (Min/Tar/Max) ----
    col = 1
    for _, header_label in IDENTITY_HEADERS:
        ws.cell(row=1, column=col, value=header_label)
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += 1

    group_start_col = col
    col_start_of = {}
    for key in selected_columns:
        width = 3  # always Min/Tar/Max, even for single-value params (Min/Max simply stay blank)
        col_start_of[key] = col
        ws.cell(row=1, column=col, value=key)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + width - 1)
        ws.cell(row=2, column=col, value="Min")
        ws.cell(row=2, column=col + 1, value="Tar")
        ws.cell(row=2, column=col + 2, value="Max")
        col += width

    total_cols = col - 1

    for r in (1, 2):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER

    # ---- data rows ----
    start_row = 3
    for i, rec in enumerate(records_sorted):
        row = start_row + i

        for c, (field_key, _) in enumerate(IDENTITY_HEADERS, start=1):
            value = rec.get(field_key, "")
            if field_key == "Ply":
                value = to_num_safe(value)
            ws.cell(row=row, column=c, value=value)

        for key in selected_columns:
            base_col = col_start_of[key]
            p = rec["params"].get(key)
            if p is None:
                continue  # leave Min/Tar/Max blank -- this file has no such spec
            ws.cell(row=row, column=base_col, value=clean_param_value(key, "Min", p["Min"]))
            ws.cell(row=row, column=base_col + 1, value=clean_param_value(key, "Tar", p["Tar"]))
            ws.cell(row=row, column=base_col + 2, value=clean_param_value(key, "Max", p["Max"]))

        for c in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = DATA_FONT
            cell.fill = DATA_FILL
            cell.border = BORDER
            cell.alignment = LEFT if c in (2, 3, 4, 5) else CENTER

    # ---- column widths / freeze panes ----
    widths = {1: 16, 2: 18, 3: 22, 4: 12, 5: 16, 6: 8, 7: 6}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for c in range(group_start_col, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 9

    ws.freeze_panes = ws.cell(row=3, column=2)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 16

    return wb


def to_num_safe(value):
    try:
        f = float(value)
        return int(f) if f.is_integer() else f
    except (ValueError, TypeError):
        return value


def save_workbook(wb, output_path):
    wb.save(output_path)
    return output_path


def workbook_to_bytes(wb):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
